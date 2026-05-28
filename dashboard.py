from tkinter import *
from openpyxl import Workbook, load_workbook
import alertemails as ae
import sendteacher as st
import os
import csv
import sys
import subprocess
import threading

from tkinter import messagebox

students = []


def scan_and_save():
    """Validate fields, launch register_face.py with the student name, then retrain and save to CSV."""
    name = entry1.get().strip()
    rollno = entry2.get().strip()
    email = entry3.get().strip()

    if not name or name == "Enter Name ..." or \
       not rollno or rollno == "Enter Roll no ..." or \
       not email or email == "Enter Email Id ...":
        messagebox.showwarning("Warning", "All fields are required before scanning!")
        return

    # Disable button to prevent double-clicks
    btn_scan.config(state=DISABLED, text="Scanning...")
    window.update()

    def run_scan():
        try:
            # 1. Capture face images
            script_path = os.path.join(os.getcwd(), "register_face.py")

            subprocess.run([sys.executable, script_path, name], check=True)

            # 2. Retrain the model with the new face
            subprocess.run(
                [sys.executable, "train_model.py"],
                check=True
            )

            # 3. Save student details to CSV
            file_name = "dataset.xlsx"
            if not os.path.exists(file_name):
                wb = Workbook()
                ws = wb.active
                ws.append(["rollno", "name", "emailid"])
                wb.save(file_name)

            wb = load_workbook(file_name)
            ws = wb.active

            ws.append([rollno, name, email, "0", "0"])
            wb.save(file_name)

            messagebox.showinfo("Success", f"Student '{name}' registered and model retrained!")
            window.destroy()

        except subprocess.CalledProcessError as e:
            messagebox.showerror("Error", f"Face scan or training failed:\n{e}")
            btn_scan.config(state=NORMAL, text="Scan & Save")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")
            btn_scan.config(state=NORMAL, text="Scan & Save")

    # Run in a thread so the Tkinter window stays responsive
    run_scan()


def register_student():
    global entry2, entry1, entry3, window, btn_scan
    window = Toplevel(root)
    window.state("zoomed")
    window.config(bg="black")

    title = Label(window, text="Register Student", font=("Arial", 20, "bold"), fg="white", bg="black")
    title.pack(pady=(100, 20))

    frame = Frame(window, bg='#1e1e1e')
    frame.place(relx=0.5, rely=0.5, anchor="center", width=420, height=380)

    Label(frame, text="Name", bg="#1e1e1e", fg="white").place(x=50, y=95)
    entry1 = Entry(frame, width=50, fg="gray34")
    entry1.insert(0, "Enter Name ...")
    entry1.place(x=50, y=120)

    Label(frame, text="Roll No", bg="#1e1e1e", fg="white").place(x=50, y=145)
    entry2 = Entry(frame, width=50, fg="gray34")
    entry2.insert(0, "Enter Roll no ...")
    entry2.place(x=50, y=170)

    Label(frame, text="Email", bg="#1e1e1e", fg="white").place(x=50, y=195)
    entry3 = Entry(frame, width=50, fg="gray34")
    entry3.insert(0, "Enter Email Id ...")
    entry3.place(x=50, y=220)

    info = Label(
        frame,
        text="Clicking 'Scan & Save' will open the camera,\ncapture 20 face images, retrain the model, then save.",
        bg="#1e1e1e", fg="gray60", font=("Arial", 9), justify=LEFT
    )
    info.place(x=50, y=260)

    btn_scan = Button(frame, text="Scan & Save", command=scan_and_save, width=15)
    btn_scan.place(x=50, y=320)

    btn_cancel = Button(frame, text="Cancel", command=window.destroy, width=10)
    btn_cancel.place(x=290, y=320)


def mark_attendance():
    """Launch the face recognition camera to mark attendance."""
    def run_recognition():
        try:
            subprocess.run([sys.executable, "recogonize_face.py"], check=True)
            messagebox.showinfo("Done", "Attendance session ended. Records saved to attendance.xlsx")
        except subprocess.CalledProcessError as e:
            messagebox.showerror("Error", f"Recognition failed:\n{e}")
        except FileNotFoundError:
            messagebox.showerror("Error", "recogonize_face.py not found. Make sure it's in the same folder.")

    # Check that the encodings file exists before launching
    if not os.path.isfile(os.path.join("encodings", "encodings.pickle")):
        messagebox.showwarning(
            "No Model Found",
            "No trained model found.\nPlease register at least one student first so the model can be built."
        )
        return

    threading.Thread(target=run_recognition, daemon=True).start()


def send_report():
    st.send_report()

def send2_report():
    st.send2_report()

#def send_emails():
   #ae.check()


def main():
    global root, title, register, mark, view1, view2, email, exit1
    root = Tk()
    root.title("Attendance Management System")
    root.state("zoomed")
    root.configure(bg="black")

    title = Label(root, text="Attendance Management System",
                  font=("Arial", 20, "bold"), fg="white", bg="black")
    title.pack(pady=(100, 20))

    register = Button(root, text="Register Student", width=20, command=register_student)
    register.pack(pady=10)

    mark = Button(root, text="Mark Attendance", width=20, command=mark_attendance)
    mark.pack(pady=10)

    view1 = Button(root, text="Send Monthly Report", width=20, command=send_report)
    view1.pack(pady=10)

    view2 = Button(root, text="Send Daily Report", width=20, command=send2_report)
    view2.pack(pady=10)

    #email = Button(root, text="Send Warning Emails", width=20, command=send_emails)
    #email.pack(pady=10)

    exit1 = Button(root, text="Exit", width=20, command=root.quit)
    exit1.pack(pady=10)

    root.mainloop()

main()
