import smtplib
import csv
import os
from email.message import EmailMessage
from tkinter import messagebox
from openpyxl import load_workbook,Workbook
from openpyxl.chart import BarChart, Reference
import matplotlib.pyplot as plt
from datetime import datetime

def send_report():
    file_name = "dataset.xlsx"
    graph_file = "overall_trend_graph.png"
    sender = "studenthelper094@gmail.com"
    app_password = "bnnh eiss lrvn yxzi"  # Remember: NO SPACES if you get auth errors
    recipient = "25pd23@psgtech.ac.in"

    if not os.path.exists(file_name):
        messagebox.showerror("Error", "dataset.xlsx not found!")
        exit()

# 2. READ ALL DATA FOR THE GRAPH
    wb = load_workbook(file_name, data_only=True)
    ws = wb.active

    dates = []
    present_counts = []

# Loop through all columns starting from Column 3 (assuming Col 1=Roll, Col 2=Name)
    for col in range(3, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col).value
    
    # Format the header (date) for the X-axis of the graph
        if isinstance(header_val, datetime):
            date_str = header_val.strftime("%m-%d")
        else:
            date_str = str(header_val)
    
        dates.append(date_str)
    
    # Count how many students were present on THIS specific day
        day_present = 0
        for row in range(2, ws.max_row + 1):
            status = str(ws.cell(row=row, column=col).value).strip().lower()
            if status == "present":
                day_present += 1
        present_counts.append(day_present)

# 3. GENERATE THE TREND GRAPH
    plt.figure(figsize=(10, 5))
# Plotting a line with markers to show the trend
    plt.plot(dates, present_counts, marker='o', linestyle='-', color='blue', linewidth=2)
    plt.fill_between(dates, present_counts, color='skyblue', alpha=0.3) # Adds a nice blue area under the line

    plt.title("Overall Attendance Trend (Total Present per Day)")
    plt.xlabel("Dates")
    plt.ylabel("Number of Students Present")
    plt.xticks(rotation=45) # Rotates dates so they don't overlap
    plt.grid(True, linestyle='--', alpha=0.6)
    plt.tight_layout()
    plt.savefig(graph_file)
    plt.close()

# 4. SEND EMAIL WITH FULL DATASET + TREND GRAPH
    msg = EmailMessage()
    msg["Subject"] = "Full Monthly Attendance Report & Trend Graph"
    msg["From"] = "Attendance Manager"
    msg["To"] = recipient
    msg.set_content("Attached is the full attendance dataset and the trend graph showing attendance for all recorded days.")

# Attach the original dataset
    with open(file_name, "rb") as f:
        msg.add_attachment(f.read(), maintype="application", 
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                       filename="dataset.xlsx")

# Attach the trend graph image
    with open(graph_file, "rb") as f:
        msg.add_attachment(f.read(), maintype="image", subtype="png", filename="attendance_trend.png")

# 5. SMTP CONNECTION
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender, app_password)
        server.send_message(msg)
        server.quit()
        messagebox.showinfo("Success", "Full Report and Trend Graph emailed successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to send email: {e}")

def send2_report():
    file_name = "dataset.xlsx" #try using marked_attendence.xlsx 

    if not os.path.exists(file_name):
        messagebox.showinfo("Important Message", "File not found!")
        return

    wb = load_workbook(file_name)
    ws = wb.active

    today = datetime.now().strftime("%Y-%m-%d")

    # -------- FIND OR CREATE TODAY COLUMN --------
    today_col = None

    for col in range(1, ws.max_column + 1):
        print(str(ws.cell(row=1,column=col).value))
        cell_value = ws.cell(row=1, column=col).value

        if isinstance(cell_value, datetime):
            cell_str = cell_value.strftime("%Y-%m-%d")
        else:
            cell_str = str(cell_value)

        if cell_str == today:
            today_col = col
            break

    if today_col is None:
        messagebox.showinfo("Important Message", "No attendance marked today!")
        return

    # -------- CREATE DAILY REPORT --------
    report_file = "daily_report.xlsx"

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["Roll No", "Name", "Status"])

    # -------- COPY DATA --------
    for row in range(2, ws.max_row + 1):
        rollno = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        status = ws.cell(row=row, column=today_col).value

        if status is None:
            status = "Absent"

        ws2.append([rollno, name, status])

    wb2.save(report_file)

    # -------- SEND EMAIL --------
    sender = "studenthelper094@gmail.com"
    app_password = "bnnh eiss lrvn yxzi"

    msg = EmailMessage()
    msg["Subject"] = "Daily Attendance Report"
    msg["From"] = "Attendance Manager"
    msg["To"] = "25pd23@psgtech.ac.in"

    msg.set_content("Please find the attached attendance report.")

    try:
        with open(report_file, "rb") as f:
            file_data = f.read()

        msg.add_attachment(
            file_data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="daily_report.xlsx"
        )
        
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender, app_password)
        server.send_message(msg)
        server.quit()

        messagebox.showinfo("Important Message", "Report emailed successfully!")

    except Exception as e:
        messagebox.showinfo("Important Message", f"Error: {e}")

send_report()