import cv2
import face_recognition
import pickle
import time
import openpyxl
from datetime import datetime
import os
import csv

def mark_excel(name):
    file_name = "attendance.xlsx"
    csv_file = "dataset.csv "

    # ------------------- GET ROLL NO FROM CSV -------------------
    rollno = None
    if os.path.exists(csv_file):
        with open(csv_file, "r") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            for row in reader:
                if row[1] == name:
                    rollno = row[0]
                    break

    if rollno is None:
        print(f"Roll number not found for {name}")
        return

    # ------------------- CREATE FILE IF NOT EXISTS -------------------
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Roll No", "Name"])
        wb.save(file_name)

    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    today = datetime.now().strftime("%Y-%m-%d")

    # ------------------- HANDLE DATE COLUMN -------------------
    headers = [cell.value for cell in ws[1]]

    if today not in headers:
        ws.cell(row=1, column=len(headers) + 1).value = today
        headers.append(today)

    col_index = headers.index(today) + 1

    # ------------------- FIND STUDENT -------------------
    found = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == rollno:
            ws.cell(row=row, column=col_index).value = "Present"
            found = True
            break

    # ------------------- ADD NEW STUDENT -------------------
    if not found:
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = rollno
        ws.cell(row=new_row, column=2).value = name
        ws.cell(row=new_row, column=col_index).value = "Present"

    wb.save(file_name)

# ------------------- LOAD ENCODINGS -------------------
with open("encodings/encodings.pickle", "rb") as f:
    data = pickle.load(f)

# ------------------- UPDATE CSV -------------------
def update_attendance(name):
    file_name = "dataset.csv"

    if not os.path.exists(file_name):
        print("dataset.csv not found!")
        return

    rows = []
    updated = False

    with open(file_name, "r") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows.append(header)

        for row in reader:
            if row[1] == name:  # name column
                try:
                    row[4] = str(int(row[4]) + 1)  # increment presentDays
                except:
                    row[4] = "1"
                updated = True
            rows.append(row)

    if updated:
        with open(file_name, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        print(f"{name} attendance updated (+1 present day)")
    else:
        print(f"{name} not found in dataset.csv")


# ------------------- CAMERA -------------------
cap = cv2.VideoCapture(0)

recognized = False  # flag to stop loop

while True:
    ret, frame = cap.read()
    if not ret:
        break

    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    boxes = face_recognition.face_locations(rgb)
    encodings = face_recognition.face_encodings(rgb, boxes)

    names = []

    for encoding in encodings:

        matches = face_recognition.compare_faces(data["encodings"], encoding)
        name = "Unknown"

        if True in matches:
            matchedIdxs = [i for (i, b) in enumerate(matches) if b]
            counts = {}

            for i in matchedIdxs:
                name = data["names"][i]
                counts[name] = counts.get(name, 0) + 1

            name = max(counts, key=counts.get)

        names.append(name)

        # ------------------- UPDATE CSV & CLOSE -------------------
        if name != "Unknown":
            update_attendance(name)
            mark_excel(name)
            print(f"{name} recognized. Closing in 2 seconds...")

            # Show result for 2 seconds
            cv2.putText(frame, f"{name} Marked", (20, 40),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            cv2.imshow("Face Recognition", frame)
            cv2.waitKey(1)

            time.sleep(2)

            recognized = True
            break

    if recognized:
        break

    # Draw boxes
    for ((top, right, bottom, left), name) in zip(boxes, names):
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.putText(frame, name, (left, top - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

    cv2.imshow("Face Recognition", frame)

    if cv2.waitKey(1) & 0xFF == 27:
        break

cap.release()
cv2.destroyAllWindows()